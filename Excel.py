import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import os

# file_name: the name of the workbook.
# returns a workbook if it exist or create one if it doesn't
def get_workbook(file_name):
    if os.path.exists(file_name):
        return openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = "Time"
        worksheet['B1'] = "Temperature (Celsius)"
        return workbook

# enters data into the worksheet
# time: the time when the temperature was taken
# temp: the temperature recorded
# worksheet: the sheet where the data will be entered
def enter_data(time, temp, worksheet):
    max_row = worksheet.max_row
    if worksheet.cell(row = max_row, column = 1).value != time: #only enter data if its different from last entr
        worksheet.cell(row = max_row + 1, column = 1).value = time
        worksheet.cell(row = max_row + 1, column = 2).value = temp

# updates the chart on the worksheet
# date: the date when the workbook is made
# worksheet: the worksheet to add this chart to
def update_chart(date, worksheet):
    chart = ScatterChart()
    chart.title = date.isoformat()
    chart.x_axis.title = worksheet["A1"].value
    chart.y_axis.title = worksheet["B1"].value
    chart.legend = None

    max_row = worksheet.max_row
    x_values = Reference(worksheet, min_col = 1, min_row = 2, max_col = 1, max_row = max_row)
    y_values = Reference(worksheet, min_col = 2, min_row = 2, max_col = 2, max_row = max_row)
    series = Series(y_values, x_values)
    chart.series.append(series)

    worksheet.add_chart(chart, "F1")
